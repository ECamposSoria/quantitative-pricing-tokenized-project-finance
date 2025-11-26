#!/usr/bin/env python3
"""
Export canonical CSVs from the Excel source of truth.

This script reads Proyecto LEO IOT.xlsx and regenerates:
- tranches.csv (using principals/rates from the params sheet)
- debt_schedule.csv (using the Debt Service sheet)
- revenue_projection.csv (using CFADS + DSRA/MRA columns from Sculpted_Waterfall)

It is intended to be run after updating the Excel source (e.g., deleveraging)
to keep CSV inputs and validators in sync.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "Proyecto LEO IOT.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "input" / "leo_iot"


def read_params_sheet(wb) -> Dict[str, float]:
    ws = wb["Params (waterfall only)"]
    params: Dict[str, float] = {}
    for key, value in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is None or value is None:
            continue
        try:
            params[str(key).strip()] = float(value)
        except (ValueError, TypeError):
            # Skip header rows or non-numeric values
            continue
    return params


def export_tranches(wb, base_rate: float, template_path: Path, output_dir: Path) -> None:
    params = read_params_sheet(wb)
    template = pd.read_csv(template_path)

    def coupon(key: str) -> float:
        return params.get(key, 0.0)

    new_rows = []
    for row in template.itertuples(index=False):
        rate_key = {
            "senior": "Rate_Senior",
            "mezzanine": "Rate_Mezz",
            "subordinated": "Rate_Sub",
        }[row.tranche_name]
        principal_key = {
            "senior": "Principal_Senior",
            "mezzanine": "Principal_Mezz",
            "subordinated": "Principal_Sub",
        }[row.tranche_name]
        principal = params.get(principal_key, row.initial_principal) * 1_000_000
        coupon_rate = coupon(rate_key)
        spread_bps = max(0.0, (coupon_rate - base_rate) * 10_000.0)
        new_rows.append(
            {
                "tranche_name": row.tranche_name,
                "priority_level": row.priority_level,
                "initial_principal": principal,
                "rate_base_type": row.rate_base_type,
                "base_rate": base_rate,
                "spread_bps": spread_bps,
                "grace_period_years": row.grace_period_years,
                "tenor_years": row.tenor_years,
                "amortization_style": row.amortization_style,
            }
        )

    df = pd.DataFrame(new_rows)
    (output_dir / "tranches.csv").write_text(df.to_csv(index=False))
    print("✓ tranches.csv exported")


def export_debt_schedule(wb, output_dir: Path) -> None:
    ws = wb["Debt Service"]
    rows: List[Dict[str, float]] = []
    cols_map = {"senior": (3, 4), "mezzanine": (5, 6), "subordinated": (7, 8)}
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=10, values_only=True):
        year = row[0]
        if year is None:
            continue
        try:
            year = int(year)
        except (ValueError, TypeError):
            # Skip header rows
            continue
        for tranche, (interest_idx, principal_idx) in cols_map.items():
            interest = row[interest_idx - 2] or 0.0
            principal = row[principal_idx - 2] or 0.0
            rows.append(
                {
                    "year": year,
                    "tranche_name": tranche,
                    "interest_due": int(round(interest * 1_000_000)),
                    "principal_due": int(round(principal * 1_000_000)),
                }
            )
    df = pd.DataFrame(rows)
    (output_dir / "debt_schedule.csv").write_text(df.to_csv(index=False))
    print("✓ debt_schedule.csv exported")


def export_revenue_projection(wb, output_dir: Path) -> None:
    ws = wb["CFADS"]
    ws_sw = wb["Sculpted_Waterfall"]
    headers = {ws_sw.cell(row=1, column=col).value: col for col in range(1, ws_sw.max_column + 1)}
    dsra_funding_col = headers.get("DSRA_Funding")
    dsra_release_col = headers.get("DSRA_Release")
    mra_funding_col = headers.get("MRA_Funding")
    mra_use_col = headers.get("MRA_Use")

    records = []
    for idx, row in enumerate(ws.iter_rows(min_row=4, max_row=18, min_col=2, max_col=9, values_only=True), start=2):
        year, revenue, opex, maintenance, wc, tax, rcapex, cfads = row
        if year is None:
            continue
        try:
            year = int(year)
        except (ValueError, TypeError):
            # Skip header rows
            continue
        dsra_funding = ws_sw.cell(row=idx, column=dsra_funding_col).value if dsra_funding_col else 0.0
        dsra_release = ws_sw.cell(row=idx, column=dsra_release_col).value if dsra_release_col else 0.0
        mra_funding = ws_sw.cell(row=idx, column=mra_funding_col).value if mra_funding_col else 0.0
        mra_use = ws_sw.cell(row=idx, column=mra_use_col).value if mra_use_col else 0.0
        cfads_calc = (revenue or 0.0) - (opex or 0.0) - (maintenance or 0.0) - (wc or 0.0) - (tax or 0.0) - (rcapex or 0.0)
        records.append(
            {
                "year": int(year),
                "revenue_gross": float(revenue or 0.0),
                "opex": float(opex or 0.0),
                "maintenance_opex": float(maintenance or 0.0),
                "working_cap_change": float(wc or 0.0),
                "tax_paid": float(tax or 0.0),
                "rcapex": float(rcapex or 0.0),
                "capex": 0.0,
                "dsra_funding": float(dsra_funding or 0.0),
                "dsra_release": float(dsra_release or 0.0),
                "mra_funding": float(mra_funding or 0.0),
                "mra_use": float(mra_use or 0.0),
                "cfads": float(cfads if cfads not in (None, "") else cfads_calc),
            }
        )
    df = pd.DataFrame(records)
    (output_dir / "revenue_projection.csv").write_text(df.to_csv(index=False))
    print("✓ revenue_projection.csv exported")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export CSV inputs from Proyecto LEO IOT.xlsx")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to Excel source.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory to write CSVs.")
    parser.add_argument(
        "--base-rate",
        type=float,
        default=None,
        help="Base rate used to infer spreads for tranches. Defaults to project_params.csv base_rate_reference.",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        raise FileNotFoundError(f"Excel not found: {args.excel}")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(args.excel, data_only=True)
    base_rate = args.base_rate
    if base_rate is None:
        project_params = pd.read_csv(DEFAULT_OUTPUT_DIR / "project_params.csv")
        base_row = project_params.loc[project_params["param"] == "base_rate_reference", "value"]
        base_rate = float(base_row.iloc[0]) if not base_row.empty else 0.0

    export_tranches(wb, base_rate, template_path=DEFAULT_OUTPUT_DIR / "tranches.csv", output_dir=args.output_dir)
    export_debt_schedule(wb, output_dir=args.output_dir)
    export_revenue_projection(wb, output_dir=args.output_dir)


if __name__ == "__main__":
    main()
