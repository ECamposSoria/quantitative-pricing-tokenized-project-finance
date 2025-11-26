#!/usr/bin/env python3
"""
Utility to create a deleveraged copy of the Excel source of truth.

The script rewrites the principal (and optional DSRA) entries in the
`Params (waterfall only)` sheet of Proyecto LEO IOT.xlsx so downstream CSV
exports and validators stay aligned with the updated capital structure.

By default it:
- Reads current principals in millions from column A/B (keys: Principal_Senior,
  Principal_Mezz, Principal_Sub, DSRA_initial).
- Applies a new total principal target using provided weights (defaults to
  60/25/15 when not specified).
- Scales DSRA_initial proportionally (can be disabled).
- Writes the result to a new Excel file without mutating the original.

Note: This does not recalculate dependent formulas inside Excel. After writing
the new workbook, use `scripts/export_csvs_from_excel.py` to regenerate CSVs
and re-run validation.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Iterable, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "Proyecto LEO IOT.xlsx"


def read_params_sheet(ws) -> Dict[str, float]:
    params: Dict[str, float] = {}
    for key, value in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is None or value is None:
            continue
        try:
            params[str(key).strip()] = float(value)
        except (ValueError, TypeError):
            # Skip header rows or non-numeric values
            continue
    return params


def write_param(ws, key: str, value: float) -> None:
    """Write to the cell where column A == key."""

    for row in ws.iter_rows(min_row=1, max_col=2):
        if str(row[0].value).strip() == key:
            row[1].value = value
            return
    raise KeyError(f"Key '{key}' not found in Params sheet")


def parse_weights(raw: str | None, defaults: Tuple[float, float, float]) -> Tuple[float, float, float]:
    if raw is None:
        return defaults
    parts = [float(x.strip()) for x in raw.split(",")]
    if len(parts) != 3:
        raise ValueError("Weights must be three comma-separated numbers, e.g., '0.6,0.25,0.15'")
    total = sum(parts)
    if total <= 0:
        raise ValueError("Weights sum must be positive")
    return tuple(p / total for p in parts)  # type: ignore[return-value]


def compute_new_principals(total_principal: float, weights: Iterable[float]) -> Dict[str, float]:
    w_senior, w_mezz, w_sub = weights
    return {
        "senior": total_principal * w_senior,
        "mezzanine": total_principal * w_mezz,
        "subordinated": total_principal * w_sub,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Deleveraging helper for Proyecto LEO IOT.xlsx")
    parser.add_argument("--input-excel", type=Path, default=DEFAULT_EXCEL, help="Path to source Excel file.")
    parser.add_argument(
        "--output-excel",
        type=Path,
        default=DEFAULT_EXCEL.with_name(DEFAULT_EXCEL.stem + "_deleveraged.xlsx"),
        help="Destination Excel path (will be overwritten).",
    )
    parser.add_argument(
        "--total-principal",
        type=float,
        default=50_000_000.0,
        help="Target total principal in USD (e.g., 50_000_000 for 50M).",
    )
    parser.add_argument(
        "--weights",
        type=str,
        default=None,
        help="Comma-separated weights for senior, mezzanine, sub (e.g., '0.6,0.25,0.15'). "
        "If omitted, uses weights from the Excel params sheet.",
    )
    parser.add_argument(
        "--scale-dsra",
        action="store_true",
        help="Scale DSRA_initial proportionally to the new total principal.",
    )
    args = parser.parse_args()

    if not args.input_excel.exists():
        raise FileNotFoundError(f"Excel not found: {args.input_excel}")

    # Load once with data_only=True to read current values
    wb_data = load_workbook(args.input_excel, data_only=True)
    if "Params (waterfall only)" not in wb_data.sheetnames:
        raise ValueError("Params (waterfall only) sheet not found in workbook.")
    params = read_params_sheet(wb_data["Params (waterfall only)"])
    wb_data.close()

    # Load again with data_only=False to preserve formulas for writing
    wb = load_workbook(args.input_excel, data_only=False)
    ws = wb["Params (waterfall only)"]

    current_principals = {
        "senior": params.get("Principal_Senior", 0.0) * 1_000_000,
        "mezzanine": params.get("Principal_Mezz", 0.0) * 1_000_000,
        "subordinated": params.get("Principal_Sub", 0.0) * 1_000_000,
    }
    current_total = sum(current_principals.values())
    default_weights = (
        params.get("Weight_Senior", 0.6),
        params.get("Weight_Mezz", 0.25),
        params.get("Weight_Sub", 0.15),
    )
    weights = parse_weights(args.weights, default_weights)
    new_principals = compute_new_principals(args.total_principal, weights)

    # Write principals in millions.
    write_param(ws, "Principal_Senior", new_principals["senior"] / 1_000_000)
    write_param(ws, "Principal_Mezz", new_principals["mezzanine"] / 1_000_000)
    write_param(ws, "Principal_Sub", new_principals["subordinated"] / 1_000_000)
    write_param(ws, "Weight_Senior", weights[0])
    write_param(ws, "Weight_Mezz", weights[1])
    write_param(ws, "Weight_Sub", weights[2])

    if args.scale_dsra and "DSRA_initial" in params:
        scaled_dsra = params["DSRA_initial"] * (args.total_principal / current_total)
        write_param(ws, "DSRA_initial", scaled_dsra)

    # Force Excel to recalculate formulas when the file is next opened
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    wb.save(args.output_excel)
    print(
        f"Deleveraged workbook written to {args.output_excel} "
        f"(total principal {args.total_principal:,.0f} USD, weights {weights})"
    )


if __name__ == "__main__":
    main()
