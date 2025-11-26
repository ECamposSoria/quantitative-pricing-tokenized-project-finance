#!/usr/bin/env python3
"""
Workaround to recalculate Excel formulas without requiring Excel application.

This script regenerates the Debt Service sheet data by calculating interest and
principal payments directly from the Params sheet, mimicking the Excel formulas.

This is necessary because openpyxl cannot evaluate formulas - it only reads
cached values. When we modify Params and save, the cached values in Debt Service
become stale (None).
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "Proyecto LEO IOT_deleveraged.xlsx"
ORIGINAL_EXCEL = ROOT / "Proyecto LEO IOT.xlsx"


def read_params_sheet(ws):
    """Read all key-value pairs from Params sheet."""
    params = {}
    for key, value in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is None or value is None:
            continue
        try:
            params[str(key).strip()] = float(value)
        except (ValueError, TypeError):
            continue
    return params


def calculate_debt_service(params: dict):
    """
    Calculate debt service schedule from params.

    Mimics the Excel Debt Service sheet formulas. Assumes:
    - Grace period: Years 1-grace_years (no principal repayment)
    - Amortization: Years grace_years+1 to tenor_years (linear principal repayment)
    - Interest: Annual, on outstanding balance
    """
    # Principals in millions
    principals = {
        "senior": params.get("Principal_Senior", 0.0),
        "mezzanine": params.get("Principal_Mezz", 0.0),
        "subordinated": params.get("Principal_Sub", 0.0),
    }
    # Interest rates (annual)
    rates = {
        "senior": params.get("Rate_Senior", 0.06),
        "mezzanine": params.get("Rate_Mezz", 0.085),
        "subordinated": params.get("Rate_Sub", 0.11),
    }

    grace_years = int(params.get("Grace_Years", 4))
    tenor_years = int(params.get("Tenor_Years", 15))
    amort_years = tenor_years - grace_years

    # Track outstanding balance for each tranche
    outstanding = {k: v for k, v in principals.items()}

    schedule = []
    for year in range(1, tenor_years + 1):
        for tranche in ["senior", "mezzanine", "subordinated"]:
            if year <= grace_years:
                # Grace period: interest only, no principal payment
                interest = outstanding[tranche] * rates[tranche]
                principal_pmt = 0.0
            else:
                # Amortization period: interest on outstanding + principal payment
                interest = outstanding[tranche] * rates[tranche]
                principal_pmt = principals[tranche] / amort_years
                outstanding[tranche] -= principal_pmt

            schedule.append({
                "year": year,
                "tranche": tranche,
                "interest": interest,
                "principal": principal_pmt,
            })

    return schedule


def write_debt_service_to_excel(wb, schedule):
    """Write calculated debt service to the Debt Service sheet."""
    ws = wb["Debt Service"]

    # Clear existing data (rows 3-17, columns C-J)
    for row in range(3, 18):
        for col in range(3, 10):
            ws.cell(row=row, column=col).value = None

    # Column mapping: Senior (C,D), Mezz (E,F), Sub (G,H)
    col_map = {
        "senior": (3, 4),
        "mezzanine": (5, 6),
        "subordinated": (7, 8),
    }

    # Write calculated values
    for entry in schedule:
        year = entry["year"]
        tranche = entry["tranche"]
        row = year + 2  # Year 1 is row 3
        int_col, prin_col = col_map[tranche]

        ws.cell(row=row, column=int_col).value = entry["interest"]
        ws.cell(row=row, column=prin_col).value = entry["principal"]

    print(f"✓ Wrote {len(schedule)} debt service entries to Debt Service sheet")


def main():
    parser = argparse.ArgumentParser(
        description="Recalculate Excel formulas for deleveraged workbook"
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL,
        help="Path to deleveraged Excel file",
    )
    parser.add_argument(
        "--original",
        type=Path,
        default=ORIGINAL_EXCEL,
        help="Path to original Excel (for copying missing params)",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        raise FileNotFoundError(f"Excel not found: {args.excel}")

    # Load deleveraged workbook
    wb = load_workbook(args.excel, data_only=False)

    # Read params from deleveraged Excel
    wb_data = load_workbook(args.excel, data_only=True)
    params = read_params_sheet(wb_data["Params (waterfall only)"])
    wb_data.close()

    # If missing critical params (rates, grace, tenor), copy from original
    critical_params = ['Rate_Senior', 'Rate_Mezz', 'Rate_Sub', 'Grace_Years', 'Tenor_Years']
    missing = [p for p in critical_params if p not in params]

    if missing and args.original.exists():
        print(f"Copying {len(missing)} missing params from original Excel: {missing}")
        wb_orig = load_workbook(args.original, data_only=True)
        orig_params = read_params_sheet(wb_orig["Params (waterfall only)"])
        wb_orig.close()

        for param in missing:
            if param in orig_params:
                params[param] = orig_params[param]
                print(f"  ✓ {param} = {orig_params[param]}")

    print(f"\nLoaded params:")
    print(f"  Principal_Senior={params.get('Principal_Senior'):.2f}M")
    print(f"  Rate_Senior={params.get('Rate_Senior', 0.0):.2%}")
    print(f"  Grace_Years={params.get('Grace_Years', 0)}")

    # Calculate debt service
    schedule = calculate_debt_service(params)

    # Write to Excel
    write_debt_service_to_excel(wb, schedule)

    # Save
    wb.save(args.excel)
    print(f"\n✓ Saved recalculated workbook to {args.excel}")


if __name__ == "__main__":
    main()
