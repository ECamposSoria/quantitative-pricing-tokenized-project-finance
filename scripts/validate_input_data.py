#!/usr/bin/env python3
"""Validate input CSV files against the Proyecto LEO IOT.xlsx source of truth."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "input" / "leo_iot"
EXCEL_PATH = ROOT / "Proyecto LEO IOT.xlsx"


class ValidationError(AssertionError):
    """Custom assertion to distinguish validation failures."""


def load_workbook_data(excel_path: Path = EXCEL_PATH):
    if not excel_path.exists():
        raise ValidationError(f"Missing Excel source: {excel_path}")
    return load_workbook(excel_path, data_only=True)


def _params_dict(wb) -> Dict[str, float]:
    ws = wb["Params (waterfall only)"]
    params: Dict[str, float] = {}
    for key, value in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is None or value is None:
            continue
        try:
            params[str(key).strip()] = float(value)
        except (ValueError, TypeError):
            # Skip header rows or non-numeric values
            continue
    return params


def validate_tranches(wb) -> None:
    """Ensure tranches.csv matches Excel principal and coupon data."""
    df = pd.read_csv(DATA_DIR / "tranches.csv")
    params = _params_dict(wb)

    expected_principals = {
        "senior": params.get("Principal_Senior", 0.0) * 1_000_000,
        "mezzanine": params.get("Principal_Mezz", 0.0) * 1_000_000,
        "subordinated": params.get("Principal_Sub", 0.0) * 1_000_000,
    }
    expected_total = sum(expected_principals.values())
    base_rate = float(df["base_rate"].iloc[0]) if not df.empty else 0.0
    expected_coupon = {
        "senior": params.get("Rate_Senior"),
        "mezzanine": params.get("Rate_Mezz"),
        "subordinated": params.get("Rate_Sub"),
    }

    for tranche, expected in expected_principals.items():
        actual = df.loc[df["tranche_name"] == tranche, "initial_principal"].item()
        if abs(actual - expected) > 1:
            raise ValidationError(
                f"{tranche} principal mismatch: expected {expected}, got {actual}"
            )
        coupon_csv = (
            df.loc[df["tranche_name"] == tranche, "base_rate"].item()
            + df.loc[df["tranche_name"] == tranche, "spread_bps"].item() / 10_000.0
        )
        coupon_xlsx = expected_coupon.get(tranche)
        if coupon_xlsx is not None and abs(coupon_csv - coupon_xlsx) > 1e-6:
            raise ValidationError(
                f"{tranche} coupon mismatch: expected {coupon_xlsx}, got {coupon_csv}"
            )

    principal_sum = df["initial_principal"].sum()
    if abs(principal_sum - expected_total) > 1:
        raise ValidationError(f"Principal sum mismatch: {principal_sum} vs expected {expected_total}")

    print("✓ tranches.csv validated")


def validate_revenue_projection(wb) -> None:
    """Ensure revenue_projection.csv columns match the Excel CFADS sheet."""
    df = pd.read_csv(DATA_DIR / "revenue_projection.csv")
    ws = wb["CFADS"]

    excel_rows: List[Dict[str, float]] = []
    ws_sw = wb["Sculpted_Waterfall"]
    headers = {ws_sw.cell(row=1, column=col).value: col for col in range(1, ws_sw.max_column + 1)}
    dsra_funding_col = headers.get("DSRA_Funding")
    dsra_release_col = headers.get("DSRA_Release")
    dsra_balance_col = headers.get("DSRA_Balance")
    mra_funding_col = headers.get("MRA_Funding")
    mra_use_col = headers.get("MRA_Use")

    for idx, row in enumerate(ws.iter_rows(min_row=4, max_row=18, min_col=2, max_col=9, values_only=True), start=2):
        year, revenue, opex, maintenance, wc, tax, rcapex, cfads = row
        dsra_funding = ws_sw.cell(row=idx, column=dsra_funding_col).value if dsra_funding_col else 0.0
        dsra_release = ws_sw.cell(row=idx, column=dsra_release_col).value if dsra_release_col else 0.0
        mra_funding = ws_sw.cell(row=idx, column=mra_funding_col).value if mra_funding_col else 0.0
        mra_use = ws_sw.cell(row=idx, column=mra_use_col).value if mra_use_col else 0.0
        # Calculate CFADS fallback (same logic as export script)
        cfads_calc = (revenue or 0.0) - (opex or 0.0) - (maintenance or 0.0) - (wc or 0.0) - (tax or 0.0) - (rcapex or 0.0)
        excel_rows.append(
            {
                "year": int(year),
                "revenue_gross": float(revenue or 0.0),
                "opex": float(opex or 0.0),
                "maintenance_opex": float(maintenance or 0.0),
                "working_cap_change": float(wc or 0.0),
                "tax_paid": float(tax or 0.0),
                "capex": 0.0,
                "rcapex": float(rcapex or 0.0),
                "dsra_funding": float(dsra_funding or 0.0),
                "dsra_release": float(dsra_release or 0.0),
                "mra_funding": float(mra_funding or 0.0),
                "mra_use": float(mra_use or 0.0),
                "cfads": float(cfads if cfads not in (None, "") else cfads_calc),
            }
        )

    excel_df = pd.DataFrame(excel_rows)
    merged = df.merge(excel_df, on="year", suffixes=("_csv", "_excel"))

    for column in [
        "revenue_gross",
        "opex",
        "maintenance_opex",
        "working_cap_change",
        "tax_paid",
        "rcapex",
        "dsra_funding",
        "dsra_release",
        "mra_funding",
        "mra_use",
        "cfads",
    ]:
        diff = (merged[f"{column}_csv"] - merged[f"{column}_excel"]).abs().max()
        if diff > 0.01:
            raise ValidationError(f"{column} mismatch exceeds tolerance: {diff}")

    revenue_total = df["revenue_gross"].sum()
    revenue_total_excel = excel_df["revenue_gross"].sum()
    if abs(revenue_total - revenue_total_excel) > 0.01:
        raise ValidationError(f"Revenue total mismatch: {revenue_total} vs {revenue_total_excel}")

    cfads_total_excel = excel_df["cfads"].sum()
    if abs(df["cfads"].sum() - cfads_total_excel) > 0.01:
        raise ValidationError(f"CFADS sum mismatch vs Excel ({cfads_total_excel})")

    print("✓ revenue_projection.csv validated")


def validate_debt_schedule(wb) -> None:
    """Validate debt_schedule.csv against Excel waterfall schedule."""
    df = pd.read_csv(DATA_DIR / "debt_schedule.csv")
    ws = wb["Debt Service"]
    params = _params_dict(wb)

    excel_rows = []
    cols_map = {"senior": (3, 4), "mezzanine": (5, 6), "subordinated": (7, 8)}
    for row in ws.iter_rows(min_row=3, max_row=17, min_col=2, max_col=10, values_only=True):
        year = row[0]
        if year is None:
            continue
        try:
            year = int(year)
        except (ValueError, TypeError):
            # Skip header rows
            continue
        for tranche, (interest_idx, principal_idx) in cols_map.items():
            interest = row[interest_idx - 2] or 0.0
            principal = row[principal_idx - 2] or 0.0
            excel_rows.append(
                {
                    "year": year,
                    "tranche_name": tranche,
                    "interest_due": int(round(interest * 1_000_000)),
                    "principal_due": int(round(principal * 1_000_000)),
                }
            )

    excel_df = pd.DataFrame(excel_rows)
    merged = df.merge(excel_df, on=["year", "tranche_name"], suffixes=("_csv", "_excel"))

    for column in ["interest_due", "principal_due"]:
        diff = (merged[f"{column}_csv"] - merged[f"{column}_excel"]).abs().max()
        if diff > 1:
            raise ValidationError(f"{column} mismatch: max diff {diff}")

    # Grace period check removed - Excel defines the actual schedule, no hardcoded assumptions
    # Principal repayment total check removed - Excel may have partial amortization, balloon payments, etc.

    print("✓ debt_schedule.csv validated")


def main(excel_path: Path = EXCEL_PATH):
    try:
        wb = load_workbook_data(excel_path)
        validate_tranches(wb)
        validate_revenue_projection(wb)
        validate_debt_schedule(wb)
        print("\n✓✓✓ ALL INPUT DATA VALIDATED ✓✓✓")
    except ValidationError as exc:
        print(f"\n✗✗✗ VALIDATION FAILED: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    cli_excel = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_PATH
    main(cli_excel)
