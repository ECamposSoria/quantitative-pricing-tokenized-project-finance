import openpyxl
import pytest

from pftoken.models import CFADSCalculator, ProjectParameters, compute_dscr_by_phase

EXCEL_PATH = "TP_Quant_Validation.xlsx"


@pytest.fixture(scope="module")
def workbook():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


@pytest.fixture(scope="module")
def project_parameters() -> ProjectParameters:
    return ProjectParameters.from_directory("data/input/leo_iot")


@pytest.fixture(scope="module")
def python_cfads(project_parameters):
    calc = CFADSCalculator.from_project_parameters(project_parameters)
    return calc.calculate_cfads_vector()


@pytest.fixture(scope="module")
def python_dscr(project_parameters, python_cfads):
    return compute_dscr_by_phase(
        python_cfads,
        project_parameters.debt_schedule,
        grace_years=project_parameters.project.grace_period_years,
        tenor_years=project_parameters.project.tenor_years,
    )


def test_excel_vs_python_cfads(workbook, python_cfads):
    ws = workbook["CFADS_Calc"]
    for year, expected in sorted(python_cfads.items()):
        excel_val = ws.cell(row=year + 1, column=9).value
        assert pytest.approx(expected, abs=0.01) == excel_val


def test_excel_vs_python_dscr(workbook, python_dscr):
    outputs = {4: 1.3559322033898304, 5: 1.5061420838568895, 11: 2.8288579514696566}
    for year, target in outputs.items():
        python_dscr_val = python_dscr[year]["value"]
        assert pytest.approx(target, abs=0.001) == python_dscr_val


def test_excel_validation_tolerance(workbook):
    ws = workbook["Comparison"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        diff_pct = float(row[3])
        assert diff_pct < 0.01, f"{row[0]} exceeds tolerance"
        assert row[4] == "PASS"
